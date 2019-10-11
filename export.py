import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive


def upload_to_drive(filename):
    gauth = GoogleAuth()
    # Try to load saved client credentials
    gauth.LoadCredentialsFile("oauthcreds.txt")
    if gauth.credentials is None:
        # Authenticate if they're not there
        gauth.LocalWebserverAuth()
    elif gauth.access_token_expired:
        # Refresh them if expired
        gauth.Refresh()
    else:
        # Initialize the saved creds
        gauth.Authorize()
    # Save the current credentials to a file
    gauth.SaveCredentialsFile("oauthcreds.txt")

    drive = GoogleDrive(gauth)

    # https://stackoverflow.com/a/22934892
    # https://stackoverflow.com/a/40236586
    drive_file = drive.CreateFile({'title': filename,
                                   "parents": [{"kind": "drive#fileLink", "id": "1bJxwrKzXzfKYRuftD0BIoOEYDobUTPUw"}]})

    # Read file and set it as a content of this instance.
    drive_file.SetContentFile("paly_from_8th.xlsx")
    drive_file.Upload()  # Upload the file.


def xlref(row, column, zero_indexed=False):
    """https://stackoverflow.com/a/37713627"""
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def set_border(ws, cell_range):
    """https://stackoverflow.com/a/34521257"""
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def fill_with_borders(ws, cell_range):
    """https://stackoverflow.com/a/43109804"""
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border


def export_spreadsheet():
    homeDir = os.path.expanduser('~')
    teams = [json.loads(open(os.path.join(homeDir, 'EMCC-2019Server/cache/teams/', team)).read()) for team in os.listdir(os.path.join(homeDir, 'EMCC-2019Server/cache/teams')) if team != '.DS_Store']

    totals = ([key for key in teams[0]['totals'].keys()], 'totals')
    l3ms = ([key for key in teams[0]['l3ms'].keys()], 'l3ms')
    SDs = ([key for key in teams[0]['SDs'].keys()], 'SDs')
    maxes = ([key for key in teams[0]['maxes'].keys()], 'maxes')
    rankings = ([key for key in teams[0]['rankings'].keys()], 'rankings')
    team_abilities = ([key for key in teams[0]['team_abilities'].keys()], 'team_abilities')
    percentages = ([key for key in teams[0]['percentages'].keys()], 'percentages')
    # sykes = ([key for key in teams[0]['sykes'].keys()], 'sykes')
    cycles = ([key for key in teams[0]['cycle_times'].keys()], 'cycle_times')

    headers = [totals, l3ms, SDs, maxes, rankings, team_abilities, percentages, cycles]  # [sykes]

    wb = openpyxl.load_workbook('paly_from_8th.xlsx')

    for sheet in wb.worksheets:
        if sheet.title != 'Team Template' and sheet.title is not 'Match Template' and sheet.title is not 'Summary Template' and sheet.title != "huh":
            wb.remove(sheet)

    raw_sheet = wb.create_sheet('Raw Export')
    raw_sheet.cell(row=1, column=1).value = 'Number'

    template = wb['huh']
    summary_sheet = wb.copy_worksheet(template)
    summary_sheet.title = 'Summary'
    summary_sheet.cell(row=1, column=1).value = 'Number'

    current_column = 2

    for header in headers:
        for key in header[0]:
            raw_sheet.cell(row=1, column=current_column).value = key
            current_column += 1

    current_row = 2
    for team in teams:
            current_column = 2
            raw_sheet.cell(row=current_row, column=1).value = team['teamNumber']

            for header in headers:
                for key in header[0]:
                    raw_sheet.cell(row=current_row, column=current_column).value = team[header[1]][key]
                    current_column += 1

            summary_sheet.cell(row=current_row, column=1).value = team['teamNumber']
            summary_sheet.cell(row=current_row, column=2).value = team['totals']['avgCargoScored']
            summary_sheet.cell(row=current_row, column=3).value = team['totals']['avgHatchesScored']
            summary_sheet.cell(row=current_row, column=4).value = team['totals']['avgPiecesScoredRocket']
            summary_sheet.cell(row=current_row, column=5).value = team['totals']['avgPiecesScoredCargoShip']
            summary_sheet.cell(row=current_row, column=6).value = team['p75s']['p75CargoScored']
            summary_sheet.cell(row=current_row, column=7).value = team['p75s']['p75HatchesScored']
            summary_sheet.cell(row=current_row, column=8).value = team['maxes']['maxCargoScored']
            summary_sheet.cell(row=current_row, column=9).value = team['maxes']['maxHatchesScored']
            summary_sheet.cell(row=current_row, column=10).value = team['totals']['avgCargoScoredSandstorm']
            summary_sheet.cell(row=current_row, column=11).value = team['totals']['avgHatchesScoredSandstorm']
            summary_sheet.cell(row=current_row, column=12).value = team['totals']['timeDefending']
            summary_sheet.cell(row=current_row, column=13).value = team['totals']['timeIncap']
            summary_sheet.cell(row=current_row, column=14).value = team['team_abilities']['placeLevel2']
            summary_sheet.cell(row=current_row, column=15).value = team['team_abilities']['placeLevel3']
            summary_sheet.cell(row=current_row, column=16).value = team['team_abilities']['startLevel2']
            summary_sheet.cell(row=current_row, column=17).value = team['team_abilities']['climbHab2']
            summary_sheet.cell(row=current_row, column=18).value = team['team_abilities']['climbHab3']
            summary_sheet.cell(row=current_row, column=19).value = team['cycle_times']['cargoOverall']
            summary_sheet.cell(row=current_row, column=20).value = team['cycle_times']['hatchOverall']
            summary_sheet.cell(row=current_row, column=21).value = team['team_abilities']['placeCargo']
            summary_sheet.cell(row=current_row, column=22).value = team['team_abilities']['placeHatch']







            template = wb['Team Template']
            team_sheet = wb.copy_worksheet(template)
            team_sheet.title = str(team['teamNumber'])

            team_sheet['A1'] = team['teamNumber']
            team_sheet['A2'] = "-"         # team['sykes']['teamName']

            team_sheet['B4'] = "-"         # team['pitscouting']['length']
            team_sheet['B5'] = "-"         # team['pitscouting']['width']
            team_sheet['B6'] = team['team_abilities']['placeLevel2']
            team_sheet['B7'] = team['team_abilities']['placeLevel3']
            team_sheet['B8'] = team['team_abilities']['startLevel2']
            team_sheet['B9'] = team['team_abilities']['climbHab2']
            team_sheet['B10'] = team['team_abilities']['climbHab3']
            team_sheet['B11'] = team['team_abilities']['placeCargo']
            team_sheet['B12'] = team['team_abilities']['placeHatch']
            team_sheet['A13'] = "-" # team['pitscouting']['drivetrain']   # This is in A because it is merged with column B

            team_sheet['B15'] = team['percentages']['percentOfTotalTeleopDefending']
            team_sheet['B16'] = team['totals']['avgTimeClimbing']

            team_sheet['E5'] = team['cycle_times']['hatchL1']
            team_sheet['E6'] = team['cycle_times']['cargoL1']
            team_sheet['E7'] = team['cycle_times']['hatchUndefended']
            team_sheet['E8'] = team['cycle_times']['hatchDefended']
            team_sheet['E9'] = team['cycle_times']['cargoUndefended']
            team_sheet['E10'] = team['cycle_times']['cargoDefended']
            team_sheet['E13'] = team['cycle_times']['p75HatchL1']
            team_sheet['E14'] = team['cycle_times']['p75CargoL1']
            team_sheet['E15'] = team['cycle_times']['p75HatchUndefended']
            team_sheet['E16'] = team['cycle_times']['p75HatchDefended']
            team_sheet['E17'] = team['cycle_times']['p75CargoUndefended']
            team_sheet['E18'] = team['cycle_times']['p75CargoDefended']

            team_sheet['H5'] = team['totals']['avgHatchTeleop']
            team_sheet['H6'] = team['totals']['avgCargoTeleop']
            team_sheet['H7'] = team['totals']['avgPiecesScoredRocket']
            team_sheet['H8'] = team['totals']['avgPiecesScoredCargoShip']
            team_sheet['H9'] = team['totals']['avgCargoScoredL1'] + team['totals']['avgHatchesScoredL1']
            team_sheet['H10'] = team['totals']['avgCargoScoredL3'] + team['totals']['avgHatchesScoredL3']

            team_sheet['H13'] = team['p75s']['p75HatchTeleop']
            team_sheet['H14'] = team['p75s']['p75CargoTeleop']
            team_sheet['H15'] = team['p75s']['p75PiecesScoredRocket']
            team_sheet['H16'] = team['p75s']['p75PiecesScoredCargoShip']
            team_sheet['H17'] = team['p75s']['p75CargoScoredL1'] + team['p75s']['p75HatchesScoredL1']
            team_sheet['H18'] = team['p75s']['p75CargoScoredL3'] + team['p75s']['p75HatchesScoredL3']

            team_sheet['K4'] = team['percentages']['percentMatchesClimbHab1']
            team_sheet['K5'] = team['percentages']['percentMatchesClimbHab2']
            team_sheet['K6'] = team['percentages']['percentMatchesClimbHab3']
            team_sheet['K7'] = team['percentages']['hab3ClimbSuccessRate']
            team_sheet['K8'] = team['percentages']['hab2ClimbSuccessRate']

            team_sheet['K14'] = team['percentages']['percentMatchesStartHab1']
            team_sheet['K15'] = team['percentages']['percentMatchesStartHab2']
            team_sheet['K16'] = team['percentages']['leftHab']
            team_sheet['K17'] = team['totals']['avgHatchesScoredSandstorm']
            team_sheet['K18'] = team['totals']['avgCargoScoredSandstorm']

            team_sheet['N5'] = team['l3ms']['l3mHatchesScored']
            team_sheet['N6'] = team['l3ms']['l3mCargoScored']
            team_sheet['N7'] = team['l3ms']['l3mPiecesScoredRocket']
            team_sheet['N8'] = team['l3ms']['l3mPiecesScoredCargoShip']
            team_sheet['N9'] = team['l3ms']['l3mCargoScoredL1'] + team['l3ms']['l3mHatchesScoredL1']
            team_sheet['N10'] = team['l3ms']['l3mCargoScoredL3'] + team['l3ms']['l3mHatchesScoredL3']

            team_sheet['N13'] = team['cycle_times']['l3mHatchL1']
            team_sheet['N15'] = team['cycle_times']['l3mCargoL1']
            team_sheet['N15'] = team['cycle_times']['l3mHatchUndefended']
            team_sheet['N15'] = team['cycle_times']['l3mHatchDefended']
            team_sheet['N15'] = team['cycle_times']['l3mCargoDefended']
            team_sheet['N15'] = team['cycle_times']['l3mCargoUndefended']

            matches_row = 22
            matches_column = 1

            matches_displayed = 0

            for timd in sorted(team['timds'], key=lambda timd_: timd_['header']['matchNumber']):
                set_border(team_sheet, xlref(matches_row, matches_column) + ":" + xlref(matches_row + 10, matches_column + 4))
                team_sheet.cell(column=matches_column, row=matches_row, value="QM " + str(timd['match_number']))
                team_sheet.cell(column=matches_column, row=matches_row).font = Font(bold=True, size=14)

                if timd['header']['isNoShow'] is True:
                    team_sheet.cell(column=matches_column + 3, row=matches_row, value="NO SHOW")
                    team_sheet.cell(column=matches_column + 3, row=matches_row).font = Font(bold=True, size=14)
                else:
                    fill_with_borders(team_sheet, xlref(matches_row + 1, matches_column) + ":" + xlref(matches_row + 7, matches_column + 1))
                    fill_with_borders(team_sheet, xlref(matches_row + 1, matches_column + 3) + ":" + xlref(matches_row + 9, matches_column + 4))
                    fill_with_borders(team_sheet, xlref(matches_row + 9, matches_column) + ":" + xlref(matches_row + 9, matches_column + 1))

                    team_sheet.cell(column=matches_column, row=matches_row + 1, value="Total Cycles")
                    team_sheet.cell(column=matches_column, row=matches_row + 2, value="TOC")
                    team_sheet.cell(column=matches_column, row=matches_row + 3, value="Climb Level")
                    team_sheet.cell(column=matches_column, row=matches_row + 4, value="Left Hab")
                    team_sheet.cell(column=matches_column, row=matches_row + 5, value="Time Incap")
                    team_sheet.cell(column=matches_column, row=matches_row + 6, value="Time Defending")
                    team_sheet.cell(column=matches_column, row=matches_row + 7, value="Time Climbing")
                    team_sheet.cell(column=matches_column, row=matches_row + 9, value="Scout")

                    team_sheet.cell(column=matches_column + 1, row=matches_row + 1, value=timd['calculated']['totalCycles'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 2, value=timd['calculated']['trueOffensiveContribution'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 3, value=timd['climb']['actualClimb'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 4, value=timd['header']['leftHab'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 5, value=timd['calculated']['timeIncap'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 6, value=timd['calculated']['timeDefending'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 7, value=timd['calculated']['timeClimbing'])
                    team_sheet.cell(column=matches_column + 1, row=matches_row + 9, value=timd['header']['scoutKey'])

                    team_sheet.cell(column=matches_column + 3, row=matches_row, value="Cycles Done")
                    team_sheet.cell(column=matches_column + 3, row=matches_row).font = Font(bold=True, size=14)
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 1, value="Hatches")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 2, value="Cargo")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 3, value="Rocket")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 4, value="Cargo Ship")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 5, value="Level 1")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 6, value="Level 2")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 7, value="Level 3")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 8, value="Hatch SS")
                    team_sheet.cell(column=matches_column + 3, row=matches_row + 9, value="Cargo SS")

                    team_sheet.cell(column=matches_column + 4, row=matches_row + 1, value=timd['calculated']['hatchesScored'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 2, value=timd['calculated']['cargoScored'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 3, value=timd['calculated']['cargoScoredRocket'] + timd['calculated']['hatchScoredRocket'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 4, value=timd['calculated']['cargoScoredCargoShip'] + timd['calculated']['hatchScoredCargoShip'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 5, value=timd['calculated']['hatchScoredLevel1'] + timd['calculated']['cargoScoredLevel1'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 6, value=timd['calculated']['hatchScoredLevel2'] + timd['calculated']['cargoScoredLevel2'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 7, value=timd['calculated']['hatchScoredLevel3'] + timd['calculated']['cargoScoredLevel3'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 8, value=timd['calculated']['hatchScoredSS'])
                    team_sheet.cell(column=matches_column + 4, row=matches_row + 9, value=timd['calculated']['cargoScoredSS'])

                matches_displayed += 1
                matches_column += 6
                if matches_displayed % 3 == 0:
                    matches_row += 12
                    matches_column = 1

            current_row += 1
            wb.save('paly_from_8th.xlsx')


if __name__ == "__main__":
    export_spreadsheet()
